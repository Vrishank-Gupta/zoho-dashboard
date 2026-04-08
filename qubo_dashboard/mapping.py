from __future__ import annotations

from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
import csv
from io import StringIO

from openpyxl import load_workbook

from .cleaning import BLANK_MARKERS, normalize_fault_code
from .config import settings


def _key(value: str | None) -> str:
    return " ".join((value or "").strip().split()).lower()


@dataclass(frozen=True, slots=True)
class MappingBundle:
    product_to_category: dict[str, str]
    fc2_to_efc: dict[str, str]


PRODUCT_MAPPING_SHEET = "Product Mapping"
EFC_MAPPING_SHEET = "EFC Mapping"
PRODUCT_MAPPING_HEADERS = ("Product", "Category")
EFC_MAPPING_HEADERS = ("EFC", "FC2")


def normalize_mapping_value(value: str | None, fallback: str) -> str:
    text = str(value or "").strip()
    return text or fallback


def normalize_product_name(product: str | None, canonical_product: str | None = None) -> str:
    raw = str(product or "").strip()
    if not raw or raw.lower() in BLANK_MARKERS:
        return "Blank Product"
    return raw


@lru_cache(maxsize=1)
def load_mappings() -> MappingBundle:
    workbook_path = settings.mapping_workbook_path
    if not workbook_path:
        return MappingBundle(product_to_category={}, fc2_to_efc={})
    path = Path(workbook_path)
    if not path.exists():
        return MappingBundle(product_to_category={}, fc2_to_efc={})

    wb = load_workbook(path, read_only=True, data_only=True)

    product_to_category: dict[str, str] = {}
    if PRODUCT_MAPPING_SHEET in wb.sheetnames:
        ws = wb[PRODUCT_MAPPING_SHEET]
        for product, category in ws.iter_rows(min_row=2, values_only=True):
            product_key = _key(str(product) if product is not None else None)
            category_text = (str(category).strip() if category is not None else "") or "Other"
            if product_key:
                product_to_category[product_key] = category_text

    fc2_to_efc: dict[str, str] = {}
    if EFC_MAPPING_SHEET in wb.sheetnames:
        ws = wb[EFC_MAPPING_SHEET]
        for efc, fc2 in ws.iter_rows(min_row=2, values_only=True):
            fc2_key = _key(str(fc2) if fc2 is not None else None)
            efc_text = (str(efc).strip() if efc is not None else "") or "Other"
            if fc2_key:
                fc2_to_efc[fc2_key] = efc_text

    return MappingBundle(product_to_category=product_to_category, fc2_to_efc=fc2_to_efc)


def clear_mapping_cache() -> None:
    load_mappings.cache_clear()


def export_product_mapping_csv() -> str:
    workbook_path = _require_workbook_path()
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(("product_name", "category"))
    if PRODUCT_MAPPING_SHEET in workbook.sheetnames:
        worksheet = workbook[PRODUCT_MAPPING_SHEET]
        rows = []
        for product_name, category in worksheet.iter_rows(min_row=2, values_only=True):
            product_text = str(product_name or "").strip()
            category_text = str(category or "").strip()
            if product_text and category_text:
                rows.append((product_text, category_text))
        for product_name, category in sorted(rows, key=lambda item: item[0].lower()):
            writer.writerow((product_name, category))
    return buffer.getvalue()


def export_efc_mapping_csv() -> str:
    workbook_path = _require_workbook_path()
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(("fault_code_level_2", "executive_fault_code"))
    if EFC_MAPPING_SHEET in workbook.sheetnames:
        worksheet = workbook[EFC_MAPPING_SHEET]
        rows = []
        for efc, fc2 in worksheet.iter_rows(min_row=2, values_only=True):
            fc2_text = str(fc2 or "").strip()
            efc_text = str(efc or "").strip()
            if fc2_text and efc_text:
                rows.append((fc2_text, efc_text))
        for fc2, efc in sorted(rows, key=lambda item: item[0].lower()):
            writer.writerow((fc2, efc))
    return buffer.getvalue()


def replace_product_mapping(rows: list[dict[str, str]]) -> None:
    workbook_path = _require_workbook_path()
    workbook = load_workbook(workbook_path)
    worksheet = _ensure_sheet(workbook, PRODUCT_MAPPING_SHEET)
    _rewrite_sheet(worksheet, PRODUCT_MAPPING_HEADERS, ((row["product_name"], row["category"]) for row in rows))
    workbook.save(workbook_path)
    clear_mapping_cache()


def replace_efc_mapping(rows: list[dict[str, str]]) -> None:
    workbook_path = _require_workbook_path()
    workbook = load_workbook(workbook_path)
    worksheet = _ensure_sheet(workbook, EFC_MAPPING_SHEET)
    _rewrite_sheet(worksheet, EFC_MAPPING_HEADERS, ((row["executive_fault_code"], row["fault_code_level_2"]) for row in rows))
    workbook.save(workbook_path)
    clear_mapping_cache()


def parse_product_mapping_csv(content: str) -> list[dict[str, str]]:
    reader = csv.DictReader(StringIO(content))
    rows: list[dict[str, str]] = []
    for row in reader:
        product_name = str(row.get("product_name") or row.get("Product") or "").strip()
        category = str(row.get("category") or row.get("Category") or "").strip()
        if not product_name or not category:
            continue
        rows.append({"product_name": product_name, "category": category})
    return rows


def parse_efc_mapping_csv(content: str) -> list[dict[str, str]]:
    reader = csv.DictReader(StringIO(content))
    rows: list[dict[str, str]] = []
    for row in reader:
        fault_code_level_2 = str(row.get("fault_code_level_2") or row.get("FC2") or "").strip()
        executive_fault_code = str(row.get("executive_fault_code") or row.get("EFC") or "").strip()
        if not fault_code_level_2 or not executive_fault_code:
            continue
        rows.append({"fault_code_level_2": fault_code_level_2, "executive_fault_code": executive_fault_code})
    return rows


def _require_workbook_path() -> Path:
    workbook_path = settings.mapping_workbook_path
    if not workbook_path:
        raise FileNotFoundError("Mapping workbook path is not configured.")
    path = Path(workbook_path)
    if not path.exists():
        raise FileNotFoundError(f"Mapping workbook not found: {path}")
    return path


def _ensure_sheet(workbook, sheet_name: str):
    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]
    return workbook.create_sheet(sheet_name)


def _rewrite_sheet(worksheet, headers: tuple[str, ...], rows) -> None:
    worksheet.delete_rows(1, worksheet.max_row or 1)
    worksheet.append(list(headers))
    for row in rows:
        worksheet.append(list(row))


def map_product_category(
    product: str | None,
    canonical_product: str | None = None,
    overrides: dict[str, str] | None = None,
) -> str:
    mappings = load_mappings()
    raw_product = str(product or "").strip()
    normalized_product = normalize_product_name(product, canonical_product)
    candidates = [
        _key(raw_product),
        _key(normalized_product),
        _key(canonical_product),
    ]
    for candidate in candidates:
        if not candidate:
            continue
        direct = (overrides or {}).get(candidate)
        if direct:
            return direct
        direct = mappings.product_to_category.get(candidate)
        if direct:
            return direct
    if normalized_product == "Blank Product":
        return "Blank Product"
    return "Other"


def map_executive_fault_code(
    fault_code_level_1: str | None,
    fault_code_level_2: str | None,
    overrides: dict[str, str] | None = None,
) -> str:
    mappings = load_mappings()
    raw_fc2 = str(fault_code_level_2 or "").strip()
    normalized_fc2 = normalize_fault_code(fault_code_level_2)
    for candidate in (_key(raw_fc2), _key(normalized_fc2)):
        if not candidate:
            continue
        direct = (overrides or {}).get(candidate)
        if direct:
            return direct
        efc = mappings.fc2_to_efc.get(candidate)
        if efc:
            return efc
    raw_fc1 = str(fault_code_level_1 or "").strip()
    normalized_fc1 = normalize_fault_code(fault_code_level_1)
    for candidate in (_key(raw_fc1), _key(normalized_fc1)):
        if not candidate:
            continue
        direct = (overrides or {}).get(candidate)
        if direct:
            return direct
        efc = mappings.fc2_to_efc.get(candidate)
        if efc:
            return efc
    return "Others"
